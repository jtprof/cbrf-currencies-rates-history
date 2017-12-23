from collections import Counter
from cbrfrates import CBRFRateByDateClient
from datetime import datetime
from datetime import timedelta
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import pandas as pd

rates = OrderedDict ()

n = 365*10+36
days = []

#Getting the data
for i in range(n):
    cbrequest = CBRFRateByDateClient(datetime.now()-timedelta(days=n-i))
    days.append(cbrequest.isodate)
    print 'Get rate by date '+ cbrequest.cbrfdate
    r = cbrequest.execute()
    for e in cbrequest.rates:
        if rates.has_key(e['CharCode']):
            if rates[e['CharCode']].has_key(cbrequest.isodate):
                pass
            else:
                rates[e['CharCode']][cbrequest.isodate] = e['Value']
        else:
            rates[e['CharCode']] = OrderedDict({cbrequest.isodate: e['Value']})


#Analize there

#Save to Excel
dest_filename = 'rates'+days[0]+'-'+days[len(days)-1]+'.xlsx'

wb = Workbook()
ws = wb.active
ws.title = "Data"

i = 2
ws.cell(column=1, row=1, value='Date')
for c in rates:
    r = ws.cell(column=i, row=1, value=c)
    i = i +1

j = 2
for d in days:
    r = ws.cell(column=1, row=j, value=datetime(int(d[0:4]), int(d[4:6]), int(d[6:8])))
    i = 2
    for c in rates:
        if rates[c].has_key(d):
            r = ws.cell(column=i, row=j, value=rates[c][d])
        i = i +1
    j = j +1

wb.save(filename = dest_filename)